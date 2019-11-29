import openpyxl as px

FILEPATH = 'C:/Users/YFC-009/Documents/Raspberry Pi/office_files/books_manager_20191126.xlsx'
SHEETNAME = 'library'

wb = px.load_workbook('FILEPATH')  # 既存のexcelファイルの読み込み
ws = wb.get_sheet_by_name('SHEETNAME')  # 処理したいシートを選択

ws['A1'].value = 123  # セルA1に数値123を入力
data = ws['B2'].vale  # セルB2の値を取得

wb.save('FILEPATH')
